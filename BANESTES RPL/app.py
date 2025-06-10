import pandas as pd
import tkinter as tk
from tkinter import filedialog
import unicodedata
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

def selecionar_arquivo():
    root = tk.Tk()
    root.withdraw()
    return filedialog.askopenfilename(
        title="Selecione o arquivo",
        filetypes=[("Arquivos Excel/CSV", "*.xlsx *.xls *.csv"), ("Todos os arquivos", "*.*")]
    )

def normalizar_texto(texto):
    if not isinstance(texto, str):
        return ""
    texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII')
    return ''.join(c for c in texto if c.isalnum() or c.isspace()).strip().lower()

def criar_nome_arquivo_saida(arquivo_original, nome_planilha):
    base, ext = os.path.splitext(arquivo_original)
    contador = 1
    while True:
        novo_nome = f"{base}_extraido_{nome_planilha}_{contador}{ext}"
        if not os.path.exists(novo_nome):
            return novo_nome
        contador += 1

def formatar_contabil(valor):
    if pd.isna(valor):
        return ""
    try:
        valor = float(valor)
        return f"{valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')  # Mantém 2 casas decimais
    except:
        return str(valor)

def calcular_saldo_total_por_dia(df, valor_saldo_anterior):
    # Converter 'Valor' para float, lidando com formatação
    def converter_valor(valor):
        try:
            if isinstance(valor, str):
                return float(valor.replace('.', '').replace(',', '.'))
            return float(valor)
        except (ValueError, TypeError):
            return 0.0  # Retorna 0 para valores não numéricos

    # Criar uma cópia do DataFrame com valores convertidos
    df = df.copy()
    df['Valor_Numerico'] = df['Valor'].apply(converter_valor)

    # Agrupar por 'Data_da_Ocorrencia' e somar os valores
    df_agrupado = df.groupby('Data_da_Ocorrencia')['Valor_Numerico'].sum().reset_index()

    # Ordenar por data para garantir a sequência correta
    df['Data_da_Ocorrencia'] = pd.to_datetime(df['Data_da_Ocorrencia'], format='%d/%m/%Y', errors='coerce')
    df_agrupado['Data_da_Ocorrencia'] = pd.to_datetime(df_agrupado['Data_da_Ocorrencia'], format='%d/%m/%Y', errors='coerce')
    df_agrupado = df_agrupado.sort_values('Data_da_Ocorrencia')

    # Calcular o saldo acumulado, começando com o valor_saldo_anterior
    saldo = converter_valor(valor_saldo_anterior) if valor_saldo_anterior else 0.0
    saldos_por_dia = [saldo]  # Inicia com o saldo do valor_saldo_anterior

    for valor in df_agrupado['Valor_Numerico']:
        saldo += valor
        saldos_por_dia.append(saldo)

    # Mapear os saldos de volta para as datas, incluindo a linha sem data
    saldo_dict = {"" : formatar_contabil(saldos_por_dia[0])}  # Saldo para a linha sem data
    saldo_dict.update(zip(df_agrupado['Data_da_Ocorrencia'].dt.strftime('%d/%m/%Y'), [formatar_contabil(s) for s in saldos_por_dia[1:]]))

    # Identificar a última data de cada mês
    df['Mes_Ano'] = df['Data_da_Ocorrencia'].dt.to_period('M')
    ultima_data_mes = df.groupby('Mes_Ano')['Data_da_Ocorrencia'].idxmax()

    # Identificar a última linha de cada data única
    ultima_linha_data = df.groupby('Data_da_Ocorrencia').tail(1).index

    # Criar série de saldos
    saldos = df['Data_da_Ocorrencia'].map(saldo_dict).fillna("")
    # Substituir o saldo pela coluna 'Valor' para a última data de cada mês
    for idx in ultima_data_mes:
        if not pd.isna(idx):
            saldos.iloc[int(idx)] = df.at[int(idx), 'Valor']

    return saldos, ultima_linha_data

def extrair_dados(arquivo):
    try:
        if arquivo.lower().endswith(('.xlsx', '.xls')):
            xls = pd.ExcelFile(arquivo)
            processar_excel(xls, arquivo)
        elif arquivo.lower().endswith('.csv'):
            processar_csv(arquivo)
        else:
            print("Formato de arquivo não suportado.")
    except Exception as e:
        print(f"Ocorreu um erro: {e}")

def processar_excel(xls, arquivo):
    for sheet_name in xls.sheet_names:
        print(f"\nProcessando planilha: {sheet_name}")
        try:
            df = pd.read_excel(arquivo, sheet_name=sheet_name, header=None)
            processar_dataframe(df, arquivo, sheet_name)
        except Exception as e:
            print(f"Erro ao processar planilha {sheet_name}: {e}")

def processar_csv(arquivo):
    print("\nProcessando arquivo CSV")
    encodings = ['utf-8', 'latin1', 'iso-8859-1', 'cp1252']
    separadores = [',', ';', '\t']
    for encoding in encodings:
        for sep in separadores:
            try:
                df = pd.read_csv(arquivo, header=None, encoding=encoding, sep=sep)
                print(f"Arquivo lido com encoding {encoding} e separador '{sep}'")
                processar_dataframe(df, arquivo, "CSV")
                return
            except:
                continue
    print("Não foi possível ler o arquivo CSV ")

def processar_dataframe(df, arquivo, nome_planilha):
    variacoes_cabecalhos = {
        'data': ['data', 'dataocorrencia', 'data_ocorrencia', 'data_da_ocorrencia', 'dataocorrência', 'data ocorrência'],
        'valor': ['valor', 'valores', 'vlr', 'val']
    }

    linha_cabecalho = None
    colunas_originais = []

    for idx, linha in df.iterrows():
        linha_normalizada = [normalizar_texto(str(cell)) for cell in linha.values]
        encontrados = {key: False for key in variacoes_cabecalhos}
        for col in linha_normalizada:
            for key, variacoes in variacoes_cabecalhos.items():
                if any(v in col for v in variacoes):
                    encontrados[key] = True
        if all(encontrados.values()):
            linha_cabecalho = idx
            colunas_originais = [str(cell).strip() for cell in linha.values]
            print(f"Cabeçalhos encontrados: {colunas_originais}")
            break

    # Procurar valor do Saldo Anterior
    valor_saldo_anterior = None
    for idx, row in df.iterrows():
        linha_normalizada = [normalizar_texto(str(cell)) for cell in row.values]
        if any('saldoanterior' in norm for norm in linha_normalizada):
            for val in row:
                try:
                    valor_saldo_anterior = float(str(val).replace('.', '').replace(',', '.'))
                    print(f"Valor do Saldo Anterior encontrado: {formatar_contabil(valor_saldo_anterior)}")
                    break
                except (ValueError, TypeError):
                    continue
            if valor_saldo_anterior is not None:
                break

    if linha_cabecalho is not None:
        print(f"Encontrados cabeçalhos na linha {linha_cabecalho + 1}")
        if valor_saldo_anterior is None:
            print("Nenhum valor de Saldo Anterior encontrado.")

        if nome_planilha == "CSV":
            df_final = pd.read_csv(arquivo, header=linha_cabecalho)
        else:
            df_final = pd.read_excel(arquivo, sheet_name=nome_planilha, header=linha_cabecalho)

        df_final = df_final.dropna(how='all')

        colunas_map = {col: normalizar_texto(col) for col in df_final.columns}
        df_final.rename(columns=colunas_map, inplace=True)

        col_data = next((col for col in df_final.columns if 'data' in col), None)
        colunas_valor = [col for col in df_final.columns if 'valor' in col]
        col_valor = colunas_valor[1] if len(colunas_valor) >= 2 else (colunas_valor[0] if colunas_valor else None)

        if not all([col_data, col_valor]):
            print("As colunas 'data' ou 'valor' não foram encontradas.")
            return

        df_final = df_final[[col_data, col_valor]]
        df_final.columns = ['Data_da_Ocorrencia', 'Valor']

        # Não remover a primeira linha nem as últimas 8 linhas
        # Não remover a última linha de cada data

        df_final['Valor'] = df_final['Valor'].apply(formatar_contabil)

        df_final['Data_da_Ocorrencia'] = pd.to_datetime(
            df_final['Data_da_Ocorrencia'], errors='coerce', dayfirst=True
        ).dt.strftime('%d/%m/%Y')

        # Adicionar a linha do Saldo Anterior no início
        if valor_saldo_anterior is not None:
            linha_saldo_anterior = pd.DataFrame({
                'Data_da_Ocorrencia': [""],
                'Valor': [formatar_contabil(valor_saldo_anterior)]
            })
            df_final = pd.concat([linha_saldo_anterior, df_final], ignore_index=True)

        # Calcular a coluna Saldo_Total com base nas somas diárias, ajustando a última data de cada mês
        saldos, ultima_linha_data = calcular_saldo_total_por_dia(df_final, valor_saldo_anterior)

        # Aplicar negrito ao valor da última linha de cada data no console
        df_final_display = df_final.copy()
        for idx in ultima_linha_data:
            if idx > 0:  # Ignorar a linha do Saldo Anterior (índice 0)
                df_final_display.at[idx, 'Valor'] = f"**{df_final.at[idx, 'Valor']}**"

        df_final['Saldo_Total'] = saldos

        print("\nDados extraídos e formatados:")
        print(df_final_display.head())
        
        # Verificar as duas últimas linhas do DataFrame resultante
        if len(df_final) >= 2:
            print("\nDuas últimas linhas do DataFrame:")
            print(df_final_display.tail(2))
        elif len(df_final) > 0:
            print("\nDataFrame resultante tem menos de 2 linhas:")
            print(df_final_display)
        else:
            print("\nDataFrame resultante está vazio.")

        nome_saida = criar_nome_arquivo_saida(arquivo, nome_planilha)
        if nome_planilha == "CSV":
            df_final.to_csv(nome_saida, index=False, encoding='utf-8')
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Dados Extraídos'

            # Mapear índices originais para linhas no Excel
            data_indices = {idx: i for i, idx in enumerate(df_final.index) if i > 0}  # Ignorar linha 0 (Saldo Anterior)

            for r_idx, row in enumerate(dataframe_to_rows(df_final, index=False, header=True), 1):
                ws.append(row)
                if r_idx > 1:  # Ignorar a linha do Saldo Anterior e o cabeçalho
                    ws[f'B{r_idx}'].number_format = '#.##0,00_-'
                    ws[f'C{r_idx}'].number_format = '#.##0,00_-'
                    # Aplicar negrito à última linha de cada data usando o mapeamento de índices
                    original_idx = df_final.index[r_idx - 2]  # Ajuste para alinhar com os dados (ignorando cabeçalho)
                    if original_idx in ultima_linha_data:
                        ws[f'B{r_idx}'].font = Font(bold=True)

            wb.save(nome_saida)

        print(f"\nNovo arquivo criado: {nome_saida}")
    else:
        print(df.head())

def main():
    arquivo = selecionar_arquivo()
    if arquivo:
        print(f"\nArquivo selecionado: {arquivo}")
        extrair_dados(arquivo)
    else:
        print("Nenhum arquivo selecionado.")

if __name__ == "__main__":
    main()