import pandas as pd
import tkinter as tk
from tkinter import filedialog
import unicodedata
import os

def selecionar_arquivo():
    root = tk.Tk()
    root.withdraw()
    return filedialog.askopenfilename(
        title="Selecione o arquivo",
        filetypes=[("Arquivos Excel/CSV", "*.xlsx *.xls *.csv"), ("Todos os arquivos", "*.*")]
    )

def normalizar_texto(texto):
    if not isinstance(texto, str):
        return ""
    texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII')
    return ''.join(c for c in texto if c.isalnum() or c.isspace()).strip().lower()

def criar_nome_arquivo_saida(arquivo_original, nome_planilha):
    base, ext = os.path.splitext(arquivo_original)
    contador = 1
    while True:
        novo_nome = f"{base}_extraido_{nome_planilha}_{contador}{ext}"
        if not os.path.exists(novo_nome):
            return novo_nome
        contador += 1

def formatar_contabil(valor):
    if pd.isna(valor):
        return ""
    try:
        valor = float(valor)
        return f"{valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    except:
        return str(valor)

def extrair_dados(arquivo):
    try:
        if arquivo.lower().endswith(('.xlsx', '.xls')):
            xls = pd.ExcelFile(arquivo)
            processar_excel(xls, arquivo)
        elif arquivo.lower().endswith('.csv'):
            processar_csv(arquivo)
        else:
            print("Formato de arquivo não suportado.")
    except Exception as e:
        print(f"Ocorreu um erro: {e}")

def processar_excel(xls, arquivo):
    for sheet_name in xls.sheet_names:
        print(f"\nProcessando planilha: {sheet_name}")
        try:
            df = pd.read_excel(arquivo, sheet_name=sheet_name, header=None)
            processar_dataframe(df, arquivo, sheet_name)
        except Exception as e:
            print(f"Erro ao processar planilha {sheet_name}: {e}")

def processar_csv(arquivo):
    print("\nProcessando arquivo CSV")
    encodings = ['utf-8', 'latin1', 'iso-8859-1', 'cp1252']
    separadores = [',', ';', '\t']
    for encoding in encodings:
        for sep in separadores:
            try:
                df = pd.read_csv(arquivo, header=None, encoding=encoding, sep=sep)
                print(f"Arquivo lido com encoding {encoding} e separador '{sep}'")
                processar_dataframe(df, arquivo, "CSV")
                return
            except:
                continue
    print("Não foi possível ler o arquivo CSV")

def processar_dataframe(df, arquivo, nome_planilha):
    variacoes_cabecalhos = {
        'data': ['data', 'dataocorrencia', 'data_ocorrencia', 'Data_da_Ocorrencia', 'dataocorrência', 'data ocorrência'],
        'valor': ['valor', 'valores', 'vlr', 'val'],
        'saldo': ['saldo', 'saldos', 'sld']
    }

    linha_cabecalho = None
    colunas_originais = []

    for idx, linha in df.iterrows():
        linha_normalizada = [normalizar_texto(str(cell)) for cell in linha.values]
        encontrados = {key: False for key in variacoes_cabecalhos}
        for col in linha_normalizada:
            for key, variacoes in variacoes_cabecalhos.items():
                if any(v in col for v in variacoes):
                    encontrados[key] = True
        if all(encontrados.values()):
            linha_cabecalho = idx
            colunas_originais = [str(cell).strip() for cell in linha.values]
            print(f"Cabeçalhos encontrados: {colunas_originais}")
            break

    if linha_cabecalho is not None:
        print(f"Encontrados cabeçalhos na linha {linha_cabecalho + 1}")

        if nome_planilha == "CSV":
            df_final = pd.read_csv(arquivo, header=linha_cabecalho)
        else:
            df_final = pd.read_excel(arquivo, sheet_name=nome_planilha, header=linha_cabecalho)

        df_final = df_final.dropna(how='all')

        colunas_map = {col: normalizar_texto(col) for col in df_final.columns}
        df_final.rename(columns=colunas_map, inplace=True)

        col_data = next((col for col in df_final.columns if 'data' in col), None)
        col_valor = next((col for col in df_final.columns if 'valor' in col), None)
        col_saldo = next((col for col in df_final.columns if 'saldo' in col), None)

        if not all([col_data, col_valor, col_saldo]):
            print("As colunas esperadas não foram encontradas")
            return

        df_final = df_final[[col_data, col_valor, col_saldo]]
        df_final.columns = ['Data_da_Ocorrencia', 'Valor', 'Saldo']

        # Mantém todas as linhas (não remove as últimas)
        # Remove apenas a linha de cabeçalho original se necessário
        df_final = df_final.iloc[1:] if linha_cabecalho == 0 else df_final

        # Formata os valores contábeis
        df_final['Valor'] = df_final['Valor'].apply(formatar_contabil)
        df_final['Saldo'] = df_final['Saldo'].apply(formatar_contabil)

        # Preserva o formato original da data (dd/mm/aaaa)
        if not pd.api.types.is_datetime64_any_dtype(df_final['Data_da_Ocorrencia']):
            try:
                # Verifica se já está no formato dd/mm/aaaa
                pd.to_datetime(df_final['Data_da_Ocorrencia'], format='%d/%m/%Y', errors='raise')
                # Se não gerou erro, já está no formato correto
            except:
                # Se não estiver no formato, tenta converter mantendo dd/mm/aaaa
                df_final['Data_da_Ocorrencia'] = pd.to_datetime(
                    df_final['Data_da_Ocorrencia'], 
                    errors='coerce'
                ).dt.strftime('%d/%m/%Y')

        print("\nDados extraídos e formatados")
        print(df_final.head())

        nome_saida = criar_nome_arquivo_saida(arquivo, nome_planilha)
        if nome_planilha == "CSV":
            df_final.to_csv(nome_saida, index=False, encoding='utf-8')
        else:
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows

            wb = Workbook()
            ws = wb.active
            ws.title = 'Dados Extraídos'

            for r_idx, row in enumerate(dataframe_to_rows(df_final, index=False, header=True), 1):
                ws.append(row)
                if r_idx > 1:
                    ws[f'B{r_idx}'].number_format = '#.##0,00_-'
                    ws[f'C{r_idx}'].number_format = '#.##0,00_-'

            wb.save(nome_saida)

        print(f"\nNovo arquivo criado: {nome_saida}")
    else:
        print("Cabeçalhos não encontrados. Visualização das primeiras linhas:")
        print(df.head())

def main():
    arquivo = selecionar_arquivo()
    if arquivo:
        print(f"\nArquivo selecionado: {arquivo}")
        extrair_dados(arquivo)
    else:
        print("Nenhum arquivo selecionado.")

if __name__ == "__main__":
    main()