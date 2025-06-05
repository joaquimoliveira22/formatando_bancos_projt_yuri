import pandas as pd
import tkinter as tk
from tkinter import filedialog
import unicodedata
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def selecionar_arquivo():
    root = tk.Tk()
    root.withdraw()
    return filedialog.askopenfilename(
        title="Selecione o arquivo",
        filetypes=[("Arquivos TXT/Excel/CSV", "*.txt *.xlsx *.xls *.csv"), ("Todos os arquivos", "*.*")]
    )

def normalizar_texto(texto):
    if not isinstance(texto, str):
        return ""
    texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII')
    return ''.join(c for c in texto if c.isalnum() or c.isspace()).strip().lower()

def criar_nome_arquivo_saida(arquivo, sufixo=""):
    base, ext = os.path.splitext(arquivo)
    novo_nome = f"{base}_{sufixo}.xlsx" if sufixo else f"{base}.xlsx"
    contador = 1
    while os.path.exists(novo_nome):
        novo_nome = f"{base}_{sufixo}_{contador}.xlsx" if sufixo else f"{base}_{contador}.xlsx"
        contador += 1
    return novo_nome

def detectar_delimitador(arquivo):
    with open(arquivo, 'r', encoding='utf-8', errors='ignore') as f:
        linha = f.readline()
        for sep in [',',';','\t','|']:
            if sep in linha:
                return sep
    return None

def carregar_dados(arquivo):
    ext = arquivo.lower().split('.')[-1]
    if ext in ['xls', 'xlsx']:
        df = pd.read_excel(arquivo)
    elif ext == 'csv':
        sep = detectar_delimitador(arquivo) or ','
        df = pd.read_csv(arquivo, sep=sep, encoding='utf-8', on_bad_lines='skip')
    elif ext == 'txt':
        sep = detectar_delimitador(arquivo) or '\t'
        df = pd.read_csv(arquivo, sep=sep, encoding='utf-8', on_bad_lines='skip')
    else:
        raise ValueError("Formato de arquivo não suportado")
    return df

def encontrar_colunas(df):
    variacoes_data = ['data', 'data_mov', 'dataocorrencia', 'data_ocorrencia', 'data movimentacao', 'data_movimentacao']
    variacoes_valor = ['valor', 'valores', 'vlr', 'val', 'montante']
    variacoes_debcred = ['deb_cred', 'debito_credito', 'debcre', 'credito_debito', 'tipo']
    
    col_data = col_valor = col_debcred = None
    
    for col in df.columns:
        col_norm = normalizar_texto(str(col))
        if any(v in col_norm for v in variacoes_data) and col_data is None:
            col_data = col
        if any(v in col_norm for v in variacoes_valor) and col_valor is None:
            col_valor = col
        if any(v in col_norm for v in variacoes_debcred) and col_debcred is None:
            col_debcred = col
    
    if col_data is None or col_valor is None:
        raise ValueError("Não foi possível encontrar as colunas 'Data_Mov' e 'Valor'")
    
    return col_data, col_valor, col_debcred

def salvar_data_valor(df, arquivo_origem):
    col_data, col_valor, col_debcred = encontrar_colunas(df)
    
    colunas_para_exportar = [col_data, col_valor]
    novos_nomes = ['Data_Mov', 'Valor']
    
    if col_debcred:
        colunas_para_exportar.append(col_debcred)
        novos_nomes.append('Deb_Cred')
    
    df_out = df[colunas_para_exportar].copy()
    df_out.columns = novos_nomes

    nome_saida = criar_nome_arquivo_saida(arquivo_origem, "data_valor")
    df_out.to_excel(nome_saida, index=False)

    if 'Deb_Cred' in df_out.columns:
        colorir_linhas(nome_saida, 'Deb_Cred')

    print(f"Arquivo com Data_Mov, Valor{', Deb_Cred' if col_debcred else ''} salvo em:\n{os.path.abspath(nome_saida)}")

def colorir_linhas(caminho_arquivo, coluna_debcred):
    wb = load_workbook(caminho_arquivo)
    ws = wb.active

    azul = PatternFill(start_color='87CEFA', end_color='87CEFA', fill_type='solid')  #cor azul
    vermelho = PatternFill(start_color='FA8072', end_color='FA8072', fill_type='solid')  #cor vermelha

    header = [cell.value for cell in ws[1]]
    idx_debcred = header.index('Deb_Cred') + 1

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        valor = row[idx_debcred - 1].value
        if isinstance(valor, str):
            if valor.strip().upper() == 'C':
                for cell in row:
                    cell.fill = azul
            elif valor.strip().upper() == 'D':
                for cell in row:
                    cell.fill = vermelho

    wb.save(caminho_arquivo)

def main():
    arquivo = selecionar_arquivo()
    if not arquivo:
        print("Nenhum arquivo selecionado.")
        return
    try:
        df = carregar_dados(arquivo)
        salvar_data_valor(df, arquivo)
    except Exception as e:
        print(f"Erro: {e}")

if __name__ == "__main__":
    main()
